"""
    Excel import submodule.
"""

# pylint: disable=R0903

from io import BytesIO
from typing import Any
from typing import Dict
from typing import Generic
from typing import List
from typing import Tuple
from typing import Type
from typing import TypeVar

from crud.bought_item import crud_bought_item
from db.base import Base
from db.models import UserModel
from exceptions import ExcelImportDataInvalidError
from exceptions import ExcelImportHeaderInvalidError
from exceptions import ExcelImportHeaderMissingError
from fastapi import HTTPException
from fastapi import UploadFile
from fastapi import status
from multilog import log
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from pydantic.fields import FieldInfo
from pydantic_core import ValidationError
from sqlalchemy.orm import Session

ModelType = TypeVar("ModelType", bound=Base)  # pylint: disable=C0103
SchemaType = TypeVar("SchemaType", bound=BaseModel)  # pylint: disable=C0103
CreateSchemaType = TypeVar("CreateSchemaType", bound=BaseModel)  # pylint: disable=C0103


class BaseExcelImport(Generic[ModelType, CreateSchemaType]):
    """Generative import class."""

    def __init__(
        self,
        db: Session,
        model: Type[ModelType],
        schema: Type[CreateSchemaType],
        db_obj_user: UserModel,
        file: UploadFile,
    ) -> None:
        """Inits the class.

        Args:
            db (Session): The database session.
            model (Type[ModelType]): The model to import.
            schema (Type[CreateSchemaType]): The schema for the import.
            db_obj_user (UserModel): The user who imports the data.
            file (UploadFile): The uploaded excel file.
        """
        self.db = db
        self.model = model
        self.schema = schema
        self.db_obj_user = db_obj_user

        self.data = file.file.read()
        self.wb: Workbook
        self.ws: Worksheet

        self._read_file()

    @staticmethod
    def _field_name_to_name_convention(field_name: str) -> str:
        """Converts the field name to the name convention (snake case to title case)"""
        return " ".join(i.capitalize() for i in str(field_name).split("_"))

    def _read_file(self) -> None:
        """Reads the data from the uploaded excel file."""

        log.info(f"Reading workbook, uploaded by user {self.db_obj_user.username!r}...")
        try:
            self.wb = load_workbook(filename=BytesIO(self.data), read_only=True, data_only=True)
            self.ws = self.wb.worksheets[0]
        except Exception as e:
            log.error(f"Failed to load workbook from user {self.db_obj_user.username}: {e}")
            raise HTTPException(status_code=422, detail=[str(e)]) from e

    # def _get_model_columns_by_name_convention(self) -> List[str]:
    #     model_cols = []
    #     for col in self.model.__table__.columns.keys():  # type: ignore
    #         model_cols.append(" ".join(i.capitalize() for i in str(col).split("_")))
    #     return model_cols

    def _get_schema_fields_by_name_convention(self) -> List[Tuple[str, FieldInfo]]:
        schema_cols = []
        for field_name, field in self.schema.model_fields.items():  # type: ignore
            schema_cols.append((self._field_name_to_name_convention(field_name), field))
        return schema_cols

    def _append_schema(
        self,
        db_obj_in: Dict[str, Any],
        db_objs_in: List[CreateSchemaType],
        skip_validation: bool = False,
    ) -> None:
        if skip_validation:
            db_objs_in.append(db_obj_in)  # type: ignore
        else:
            db_objs_in.append(self.schema(**db_obj_in))

    def _create(self, db_objs_in: List[CreateSchemaType]) -> List[ModelType]:
        """Creates the data in the database.

        Args:
            db_objs_in (List[CreateSchemaType]): The data to create.

        Returns:
            List[ModelType]: The created data.
        """
        items = []
        for obj_in in db_objs_in:
            items.append(
                crud_bought_item.create(
                    self.db, db_obj_user=self.db_obj_user, obj_in=obj_in  # type:ignore
                )
            )
        return items

    def get_header_row(self) -> int:
        create_schema_fields = self._get_schema_fields_by_name_convention()
        create_schema_fields_names = [n for n, _ in create_schema_fields]
        empty_row_count = 0

        for row_index in range(1, self.ws.max_row + 1):
            # Abort after 100 empty rows
            # Sometimes ws.max_row doesn't return the correct value and the loop
            # would run forever...
            if empty_row_count > 100:
                break

            header_candidate = []
            for col_index in range(1, self.ws.max_column + 1):
                cell_value = self.ws.cell(row_index, col_index).value
                log.debug(f"Reading header candidate in row {row_index} and col {col_index}...")
                if cell_value is not None:
                    header_candidate.append(cell_value)

            if not header_candidate:
                empty_row_count += 1
            else:
                empty_row_count = 0
                # Basic check if there are items of the header in the line
                if set(header_candidate).intersection(set(create_schema_fields_names)):
                    # Detail check, if every required header is present
                    for field_name, field in create_schema_fields:
                        if field.is_required() and field_name not in header_candidate:
                            log.warning(f"Cannot find required field in EXCEL column: {field_name}")
                            raise ExcelImportHeaderInvalidError(
                                f"Header is invalid: Missing column '{field_name}'",
                            )
                    log.debug(f"Import file header row is {row_index}")
                    return row_index

        log.warning(f"Failed to read header data from workbook, uploaded by user {self.db_obj_user.username!r}.")
        raise ExcelImportHeaderMissingError("Header is missing")

    def get_data_as_create_schema(self, skip_validation: bool = False) -> List[CreateSchemaType]:
        db_objs_in: List[CreateSchemaType] = []
        warnings: List[dict] = []
        header_row = self.get_header_row()

        for row_index in range(header_row + 1, self.ws.max_row + 1):
            # Sometimes ws.max_row doesn't work correct and the
            # row is empty. With this we check if there are values
            # in at least one cell in the current row.
            # If there are no values in any cell of the current row,
            # the import stops with the previous row.
            row_is_valid = False
            log.debug(f"Reading data from row {row_index}...")

            db_obj_in = {}
            for col_index in range(1, self.ws.max_column + 1):
                key = "_".join(i.lower() for i in str(self.ws.cell(row=header_row, column=col_index).value).split(" "))
                value = self.ws.cell(row_index, col_index).value

                if value is not None:
                    row_is_valid = True
                db_obj_in[key] = value

            if not row_is_valid:
                log.debug(f"Stopping reading file. Row {row_index} contains no data.")
                break

            try:
                self._append_schema(db_obj_in=db_obj_in, db_objs_in=db_objs_in, skip_validation=skip_validation)
            except ValidationError as error:
                warnings.append({"row": row_index, "errors": error.errors()})

        if warnings:
            log.warning(f"Failed to read data from workbook, uploaded by user {self.db_obj_user.username!r}.")
            raise HTTPException(status_code=422, detail=warnings)

        return db_objs_in

    def batch_create(self) -> List[ModelType]:
        """Main function for loading the data from the excel file into the database."""
        db_objs_in = self.get_data_as_create_schema()
        db_objs = self._create(db_objs_in)
        return db_objs
