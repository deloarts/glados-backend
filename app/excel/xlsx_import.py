"""
    Excel import submodule.
"""

# pylint: disable=R0903

from io import BytesIO
from typing import Generic, List, Type, TypeVar

from crud import crud_bought_item
from db.base import Base
from fastapi import HTTPException, UploadFile
from fastapi.encoders import jsonable_encoder
from models import model_user
from multilog import log
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from pydantic.error_wrappers import ValidationError
from sqlalchemy.orm import Session

ModelType = TypeVar("ModelType", bound=Base)  # pylint: disable=C0103
SchemaType = TypeVar("SchemaType", bound=BaseModel)  # pylint: disable=C0103
CreateSchemaType = TypeVar("CreateSchemaType", bound=BaseModel)  # pylint: disable=C0103


class ImportExcel(Generic[ModelType, CreateSchemaType]):
    """Generative import class."""

    def __init__(
        self,
        db: Session,
        model: Type[ModelType],
        schema: Type[CreateSchemaType],
        db_obj_user: model_user.User,
        file: UploadFile,
    ) -> None:
        """Inits the class.

        Args:
            db (Session): The database session.
            model (Type[ModelType]): The model to import.
            schema (Type[CreateSchemaType]): The schema for the import.
            db_obj_user (model_user.User): The user who imports the data.
            file (UploadFile): The uploaded excel file.
        """
        self.db = db
        self.model = model
        self.schema = schema
        self.db_obj_user = db_obj_user

        self.data = file.file.read()
        self.wb: Workbook
        self.ws: Worksheet

    @staticmethod
    def _apply_schema(data, schema) -> List[dict]:
        """Applies the schema to the data.

        Args:
            data (List[ModelType]): The data models to convert to a list of dicts.
            schema (Type[SchemaType]): The schema for the columns.

        Returns:
            List[dict]: The converted data.
        """
        return [jsonable_encoder(schema(**item.__dict__)) for item in data]

    def _read_file(self) -> None:
        """Reads the data from the uploaded excel file."""

        log.info(f"Reading workbook, uploaded by user {self.db_obj_user.username!r}...")
        try:
            self.wb = load_workbook(
                filename=BytesIO(self.data), read_only=True, data_only=True
            )
            self.ws = self.wb.worksheets[0]
        except Exception as e:
            log.error(
                f"Failed to load workbook from user {self.db_obj_user.username}: {e}"
            )
            raise HTTPException(status_code=422, detail=[str(e)])

    def _get_header_row(self) -> int:
        """Return the header row with EXCEL index (index starts by 1)

        Raises:
            HTTPException: Raises status code 422 when the header isn't found or valid.

        Returns:
            int: The header row index.
        """
        db_columns = []
        empty_row_count = 0

        for col in self.model.__table__.columns.keys():  # type: ignore
            db_columns.append(" ".join(i.capitalize() for i in str(col).split("_")))

        for row_index in range(1, self.ws.max_row + 1):
            # Abort after 100 empty rows
            # Sometimes ws.max_row doesn't return the correct value and the loop
            # would run forever...
            if empty_row_count > 100:
                break

            header_candidate = []
            for col_index in range(1, self.ws.max_column + 1):
                cell_value = self.ws.cell(row_index, col_index).value
                log.debug(
                    f"Reading header candidate in row {row_index} and col {col_index}..."
                )
                if cell_value is not None:
                    header_candidate.append(cell_value)

            if not header_candidate:
                empty_row_count += 1
            else:
                empty_row_count = 0
                if set(header_candidate) <= set(db_columns):
                    log.debug(f"Import file header row is {row_index}")
                    return row_index

        log.warning(
            f"Failed to read header data from workbook, uploaded by user {self.db_obj_user.username!r}."
        )
        raise HTTPException(status_code=422, detail=["Header missing or invalid."])

    def _read_data(self) -> List[CreateSchemaType]:
        """Reads the data from the worksheet.

        Raises:
            HTTPException: Raised if the worksheet content doesn't fit the schema.

        Returns:
            List[CreateSchemaType]: The data as schema.
        """
        db_objs_in: List[CreateSchemaType] = []
        warnings: List[str] = []
        header_row = self._get_header_row()

        for row_index in range(header_row + 1, self.ws.max_row + 1):
            # Sometimes ws.max_row doesn't work correct and the
            # row is empty. With this we check if there are values
            # in at least one cell in the current row.
            # If there are no values in any cell of the current row,
            # the import stops with the previous row.
            row_is_valid = False
            log.debug(f"Reading data from row {row_index}...")

            db_obj_in = {}
            for col_index in range(1, self.ws.max_column + 1):
                key = "_".join(
                    i.lower()
                    for i in str(
                        self.ws.cell(row=header_row, column=col_index).value
                    ).split(" ")
                )
                value = self.ws.cell(row_index, col_index).value

                if value is not None:
                    row_is_valid = True
                db_obj_in[key] = value

            if not row_is_valid:
                log.debug(f"Stopping reading file. Row {row_index} contains no data.")
                break

            try:
                db_objs_in.append(self.schema(**db_obj_in))
            except ValidationError as error:
                warnings.append(f"Error in row #{row_index}: {error}")

        if warnings:
            log.warning(
                f"Failed to read data from workbook, uploaded by user {self.db_obj_user.username!r}."
            )
            raise HTTPException(status_code=422, detail=warnings)

        return db_objs_in

    def _create(self, db_objs_in: List[CreateSchemaType]) -> List[ModelType]:
        """Creates the data in the database.

        Args:
            db_objs_in (List[CreateSchemaType]): The data to create.

        Returns:
            List[ModelType]: The created data.
        """
        items = []
        for obj_in in db_objs_in:
            items.append(
                crud_bought_item.bought_item.create(
                    self.db, db_obj_user=self.db_obj_user, obj_in=obj_in  # type:ignore
                )
            )
        return items

    def load(self) -> List[ModelType]:
        """Main function for loading the data from the excel file into the database."""
        self._read_file()
        db_objs_in = self._read_data()
        db_objs = self._create(db_objs_in)
        return db_objs
