"""
    EXCEL Export submodule.
"""

# pylint: disable=R0903

from datetime import datetime
from pathlib import Path
from typing import Generic
from typing import List
from typing import Type
from typing import TypeVar

from const import TEMP
from db.base import Base
from excel.style import style_worksheet
from fastapi.encoders import jsonable_encoder
from multilog import log
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

ModelType = TypeVar("ModelType", bound=Base)  # pylint: disable=C0103
SchemaType = TypeVar("SchemaType", bound=BaseModel)  # pylint: disable=C0103


class BaseExcelExport(Generic[ModelType, SchemaType]):
    """Generic excel export class."""

    def __init__(self, schema: Type[SchemaType]) -> None:
        """Inits the class.

        Args:
            data (List[ModelType]): The data models to convert to an xlsx file.
            schema (Type[SchemaType]): The schema for the columns.
        """
        self.schema = schema
        self.data: List[dict] = []  # self._apply_schema(data, schema)
        self.name = datetime.now().strftime("%Y%m%d_%H%M%S")

        self.wb: Workbook = Workbook()
        self.ws: Worksheet = self.wb.active  # type: ignore
        assert isinstance(self.ws, Worksheet)
        self.ws.title = self.name

    def import_data_by_model(self, data: List[ModelType]) -> None:
        self.data = [jsonable_encoder(self.schema(**item.__dict__)) for item in data]

    def import_data_by_dict(self, data: List[dict]) -> None:
        self.data = [jsonable_encoder(self.schema(**item)) for item in data]

    def _write_header(self) -> None:
        """
        Creates the header from the model column names.
        Converts underscores to spaces, Capitalizes all words.
        """
        for col_index, col in enumerate(self.data[0]):
            db_col_name = " ".join(i.capitalize() for i in str(col).split("_"))
            self.ws.cell(1, col_index + 1, db_col_name)

    def _write_data(self) -> None:
        """Writes the data to the worksheet."""
        for row_index, row in enumerate(self.data):
            for col_index, col in enumerate(row):
                self.ws.cell(row_index + 2, col_index + 1, row[col])

    def _style_worksheet(self) -> None:
        """Styles the worksheet by the config file."""
        style_worksheet(self.ws)

    def save(self) -> Path:
        """Saves the created excel file to the temp folder and returns the path."""
        if len(self.data) == 0:
            raise Exception(
                "Cannot save XLSX file: No data. "
                "Data must be imported first by calling `import_data_by_model` or `import_data_by_dict`."
            )

        self._write_header()
        self._write_data()
        self._style_worksheet()

        path = Path(TEMP, f"{self.name}.xlsx")
        self.wb.save(str(path))
        log.info(f"Saved EXCEL file at {str(path)!r}.")
        return path
