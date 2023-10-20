"""
    EXCEL Style submodule.
"""

from config import cfg
from multilog import log
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def style_worksheet(worksheet: Worksheet) -> None:
    """Styles the worksheet by the config.yml.

    Args:
        worksheet (Worksheet): The worksheet to style.
    """

    # Note: I am getting a typing error with Pylance that appears incorrect and is
    # not reported as an error by mypy --strict. The error occurs everywhere in this
    # file where type: ignore is set.

    for column_cells in worksheet.columns:
        # Set format, font and size
        for index, cell in enumerate(column_cells):  # type: ignore
            if index > cfg.excel.data_row - 1:
                color = (
                    cfg.excel.style.data_color_1
                    if index % 2 == 0
                    else cfg.excel.style.data_color_2
                )
                bg_color = (
                    cfg.excel.style.data_bg_color_1
                    if index % 2 == 0
                    else cfg.excel.style.data_bg_color_2
                )
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
            else:
                color = None
            cell.number_format = "@"
            cell.font = Font(
                name=cfg.excel.style.font, size=cfg.excel.style.size, color=color
            )
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Set font and height for the header row
        if isinstance(cfg.excel.header_row, int):
            worksheet.row_dimensions[  # type: ignore
                cfg.excel.header_row + 1
            ].height = 20  # type:ignore
            column_cells[cfg.excel.header_row].font = Font(  # type: ignore
                name=cfg.excel.style.font,
                size=cfg.excel.style.size,
                bold=True,
                color=cfg.excel.style.header_color,
            )
            column_cells[cfg.excel.header_row].fill = PatternFill(  # type: ignore
                start_color=cfg.excel.style.header_bg_color,
                end_color=cfg.excel.style.header_bg_color,
                fill_type="solid",
            )
            column_cells[cfg.excel.header_row].alignment = Alignment(  # type: ignore
                horizontal="center", vertical="center"
            )

        # Set cell width
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)  # type: ignore
        worksheet.column_dimensions[  # type:ignore
            column_cells[0].column_letter  # type: ignore
        ].width = (  # type:ignore
            length if length > 2 else 2
        )

    log.info(f"Styled worksheet {worksheet.title!r}.")
